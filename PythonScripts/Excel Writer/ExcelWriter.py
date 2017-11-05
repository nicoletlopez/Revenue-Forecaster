from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
#Templaate Name
template = load_workbook('Default.xlsx')

#Sheet Name
SheetName = template['YR2015']

#Project Name
SheetName['B1'] = 'Project Name'

#Project Year
SheetName['B2'] = 'Jan-Dec, 2017'

#Forecast for what year? Tentative Value is 2017
yearForecast = 2017

#Change Dates in Excel for Excel Output
Actualyear = 'Actual ' + str(yearForecast) #Actual Year
PreviousYear = 'Actual ' + str(yearForecast-1) #Previous Year Actual

#Insert Values to their designated row/column for Excel Output Report
for i in range(8,64, 4):
    SheetName.cell(row=i, column=3, value=Actualyear)
for i in range(10, 66, 4):
    SheetName.cell(row=i, column=3, value=PreviousYear)

#Budget Row has color red font
for x in range(9, 65, 4):
    for y in range(4,73):
        a = SheetName.cell(row=x, column=y)
        a.font = Font(color="ff0000")

SheetName['D9'] = 600
#Excel File Output with Forecast
template.save(str(yearForecast) + ' Room Segmentation.xlsx')
