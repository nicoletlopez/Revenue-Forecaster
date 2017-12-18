from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string


class ExcelWriter(object):
    def __init__(self, project_name='ProjectName', month='January', year='2017', file_path='ForecastOutput.xltx'):
        self.workbook = load_workbook(file_path)
        self.workbook.template = False
        self.worksheet = self.workbook['Forecast']
        self.default_value = 'No Forecast Yet'
        self.project_name = project_name
        self.month_year = month + ", " + year
        self.worksheet['C1'] = self.project_name
        self.worksheet['C2'] = self.month_year

    # Insert one value to any cell in the Individual/Group Forecast/Actual table.

    def insert_individual_forecast_value(self, forecast_value, sub_segment, metric):
        sub_segment_index = self.__check_individual_forecast_subsegment(sub_segment)
        metric_index = self.__check_individual_forecast_metric(metric)
        forecast_cell = self.worksheet.cell(row=metric_index, column=sub_segment_index)
        forecast_cell.value = forecast_value

    def insert_group_forecast_value(self, forecast_value, sub_segment, metric):
        sub_segment_index = self.__check_group_forecast_subsegment(sub_segment)
        metric_index = self.__check_group_forecast_metric(metric)
        forecast_cell = self.worksheet.cell(row=metric_index, column=sub_segment_index)
        forecast_cell.value = forecast_value

    def insert_individual_actual_value(self, actual_value, sub_segment, metric):
        sub_segment_index = self.__check_individual_actual_subsegment(sub_segment)
        metric_index = self.__check_individual_actual_metric(metric)
        actual_cell = self.worksheet.cell(row=metric_index, column=sub_segment_index)
        actual_cell.value = actual_value

    def insert_group_actual_value(self, actual_value, sub_segment, metric):
        sub_segment_index = self.__check_group_actual_subsegment(sub_segment)
        metric_index = self.__check_group_actual_metric(metric)
        actual_cell = self.worksheet.cell(row=metric_index, column=sub_segment_index)
        actual_cell.value = actual_value

    # Total tables don't need subsegment index since there is only one column

    def insert_total_forecast_value(self, actual_value, metric):
        metric_index = self.__check_individual_actual_metric(metric)
        actual_cell = self.worksheet.cell(row=metric_index, column=20)
        actual_cell.value = actual_value

    def insert_total_actual_value(self, actual_value, metric):
        metric_index = self.__check_group_actual_metric(metric)
        actual_cell = self.worksheet.cell(row=metric_index, column=27)
        actual_cell.value = actual_value

    # Save Workbook

    def save_file(self):
        title = self.project_name + ' - Forecast Report, ' + self.month_year + '.xlsx'
        self.workbook.save(title)

    # Check for Individual/Group Forecast Subsegment/Metric Index

    def __check_individual_forecast_subsegment(self, sub_segment):
        for sub_segment_range in self.worksheet.iter_rows(min_row=5, max_row=5, min_col=6, max_col=22):
            for cell in sub_segment_range:
                if cell.value == sub_segment:
                    c = cell.column
                    r = cell.row
                    cr = c + str(r)
                    xy = coordinate_from_string(cr)
                    col = column_index_from_string(xy[0])
                    return col
        error = "Error: wrong segment"
        print(error)

    def __check_group_forecast_subsegment(self, sub_segment):
        for sub_segment_range in self.worksheet.iter_rows(min_row=5, max_row=5, min_col=26, max_col=41):
            for index, cell in enumerate(sub_segment_range):
                if cell.value == sub_segment:
                    c = cell.column
                    r = cell.row
                    cr = c + str(r)
                    xy = coordinate_from_string(cr)
                    col = column_index_from_string(xy[0])
                    return col
        error = "Error: wrong segment"
        print(error)

    def __check_individual_forecast_metric(self, metric):
        for metric_range in self.worksheet.iter_rows(min_row=6, max_row=12, min_col=5, max_col=5):
            for cell in metric_range:
                if cell.value == metric:
                    return cell.row

    def __check_group_forecast_metric(self, metric):
        for metric_range in self.worksheet.iter_rows(min_row=6, max_row=12, min_col=25, max_col=25):
            for index, cell in enumerate(metric_range):
                if cell.value == metric:
                    return cell.row

    # Check for Individual/Group Actual Subsegment/Metric Index

    def __check_individual_actual_subsegment(self, sub_segment):
        for sub_segment_range in self.worksheet.iter_rows(min_row=16, max_row=16, min_col=6, max_col=22):
            for index, cell in enumerate(sub_segment_range):
                if cell.value == sub_segment:
                    c = cell.column
                    r = cell.row
                    cr = c + str(r)
                    xy = coordinate_from_string(cr)
                    col = column_index_from_string(xy[0])
                    return col
        error = "Error: wrong segment"
        print(error)

    def __check_group_actual_subsegment(self, sub_segment):
        for sub_segment_range in self.worksheet.iter_rows(min_row=16, max_row=16, min_col=26, max_col=41):
            for index, cell in enumerate(sub_segment_range):
                if cell.value == sub_segment:
                    c = cell.column
                    r = cell.row
                    cr = c + str(r)
                    xy = coordinate_from_string(cr)
                    col = column_index_from_string(xy[0])
                    return col
        error = "Error: wrong segment"
        print(error)

    def __check_individual_actual_metric(self, metric):
        for metric_range in self.worksheet.iter_rows(min_row=17, max_row=23, min_col=5, max_col=5):
            for index, cell in enumerate(metric_range):
                if cell.value == metric:
                    return cell.row

    def __check_group_actual_metric(self, metric):
        for metric_range in self.worksheet.iter_rows(min_row=17, max_row=23, min_col=25, max_col=25):
            for index, cell in enumerate(metric_range):
                if cell.value == metric:
                    return cell.row

    # Check for Subsegment/Metric Index of Total Actual/Forecasted

    def __check_total_forecast_metric(self, metric):
        for metric_range in self.worksheet.iter_rows(min_row=28, max_row=34, min_col=17, max_col=17):
            for index, cell in enumerate(metric_range):
                if cell.value == metric:
                    return cell.row

    def __check_total_actual_metric(self, metric):
        for metric_range in self.worksheet.iter_rows(min_row=28, max_row=34, min_col=25, max_col=25):
            for index, cell in enumerate(metric_range):
                if cell.value == metric:
                    return cell.row
