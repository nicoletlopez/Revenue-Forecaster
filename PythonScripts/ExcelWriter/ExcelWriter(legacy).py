from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

# Templaate Name
wb = load_workbook('Book1.xltx')

wb.template = False

# Sheet Name
worksheet = wb['Sheet1']

# Project Name
worksheet['C1'] = 'Project Name'

# Project Year
worksheet['C2'] = 'January, 2017'

default_value = 'No Forecast Yet'

for row in worksheet.iter_rows(min_row=6, min_col=3, max_row=6, max_col=19):
    for x, cell in enumerate(row):
        cell.value = default_value
        print(x)

wb.save('Book2.xlsx')

# #Forecast for what year? Tentative Value is 2017
# yearForecast = 2017
#
# #Change Dates in Excel for Excel Output
# Actualyear = 'Actual ' + str(yearForecast) #Actual Year
# PreviousYear = 'Actual ' + str(yearForecast-1) #Previous Year Actual
#
# #Insert Values to their designated row/column for Excel Output Report
# for i in range(8,64, 4):
#     worksheet.cell(row=i, column=3, value=Actualyear)
# for i in range(10, 66, 4):
#     worksheet.cell(row=i, column=3, value=PreviousYear)
#
# #Budget Row has color red font
# for x in range(9, 65, 4):
#     for y in range(4,73):
#         a = worksheet.cell(row=x, column=y)
#         a.font = Font(color="ff0000")
#
# worksheet['D9'] = 600
# #Excel File Output with Forecast
#
# wb.save('document.xlsx'